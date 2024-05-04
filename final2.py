from ultralytics import RTDETR
import cv2
import io
from PIL import Image
from flask import Flask, request, redirect, url_for,send_file
from flask_cors import CORS
import zipfile
from pdf2image import convert_from_path
from werkzeug.utils import secure_filename


import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.patches import Patch
import io
from PIL import Image, ImageDraw
import numpy as np
import csv
import pandas as pd

from torchvision import transforms

from transformers import AutoModelForObjectDetection
import torch




import shutil
import re
import pytesseract
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
# from openpyxl.drawing.image import Image as pyxl_image
from openpyxl.utils import get_column_letter
import os


app = Flask(__name__)

CORS(app)
detection_model= RTDETR('table_detection.pt')
@app.route('/', methods=['GET'])

def hello():
    return 'Hello, World!'

@app.route('/upload', methods=['GET', 'POST'])

def table_image():
    if request.method =='POST':
        img_file = request.files['the_file']
        
        if img_file.filename.endswith('.pdf'):
            pdf_file_name = secure_filename(img_file.filename)
            img_file.save(pdf_file_name)
            img_file = convert_from_path(pdf_file_name)
            img_file[0].save('image_got.jpg','JPEG')
            img_file = 'image_got.jpg'
    


        coordinates_lis = table_detect(img_file)
        full_image = Image.open(img_file)
        with zipfile.ZipFile('output.zip','w') as zip_file:
            pass

        for idx,val in enumerate(coordinates_lis):
            img_crop = full_image.crop(val)
            final_output(img_crop)
            with zipfile.ZipFile('output.zip','a') as zip_file:
                zip_file.write('output_final.xlsx',arcname=f"output{idx}.xlsx")

        # img_io = io.BytesIO()
        # cropped_img.save(img_io, 'PNG')
        # img_io.seek(0)
        return send_file('output.zip', as_attachment=True)
        
    return 'noo'


def table_detect(file):
    lis = []
    img = Image.open(file)
    results = detection_model(img)
    for result in results:
        boxes = result.boxes  # Boxes object for bounding box outputs
        for box in boxes.xyxy:
            lis2 = []
            for coordinate in box:
                lis2.append(int(coordinate.item()))
            lis.append(lis2)
    return lis




device = "cuda" if torch.cuda.is_available() else "cpu"


class MaxResize(object):
    def __init__(self, max_size=800):
        self.max_size = max_size

    def __call__(self, image):
        width, height = image.size
        current_max_size = max(width, height)
        scale = self.max_size / current_max_size
        resized_image = image.resize((int(round(scale*width)), int(round(scale*height))))
        
        return resized_image



structure_transform = transforms.Compose([
    MaxResize(1000),
    transforms.ToTensor(),
    transforms.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
])

structure_model = AutoModelForObjectDetection.from_pretrained("microsoft/table-transformer-structure-recognition-v1.1-all").to(device)

def recognize_table(image):
    # prepare image for the model
    # pixel_values = structure_processor(images=image, return_tensors="pt").pixel_values
    pixel_values = structure_transform(image).unsqueeze(0).to(device)

    # forward pass
    with torch.no_grad():
        outputs = structure_model(pixel_values)

    # postprocess to get individual elements
    id2label = structure_model.config.id2label
    id2label[len(structure_model.config.id2label)] = "no object"
    cells = outputs_to_objects(outputs, image.size, id2label)

    # visualize cells on cropped table

    draw = ImageDraw.Draw(image)

    for cell in cells:
        draw.rectangle(cell["bbox"], outline="red")
        
    return image, cells


def get_cell_coordinates_by_row(table_data):
    # Extract rows and columns
    rows = [entry for entry in table_data if entry['label'] == 'table row']
    columns = [entry for entry in table_data if entry['label'] == 'table column']

    # Sort rows and columns by their Y and X coordinates, respectively
    rows.sort(key=lambda x: x['bbox'][1])
    columns.sort(key=lambda x: x['bbox'][0])

    # Function to find cell coordinates
    def find_cell_coordinates(row, column):
        cell_bbox = [column['bbox'][0], row['bbox'][1], column['bbox'][2], row['bbox'][3]]
        return cell_bbox

    # Generate cell coordinates and count cells in each row
    cell_coordinates = []

    for row in rows:
        row_cells = []
        for column in columns:
            cell_bbox = find_cell_coordinates(row, column)
            row_cells.append({'column': column['bbox'], 'cell': cell_bbox})

        # Sort cells in the row by X coordinate
        row_cells.sort(key=lambda x: x['column'][0])

        # Append row information to cell_coordinates
        cell_coordinates.append({'row': row['bbox'], 'cells': row_cells, 'cell_count': len(row_cells)})

    # Sort rows from top to bottom
    cell_coordinates.sort(key=lambda x: x['row'][1])

    return cell_coordinates



def box_cxcywh_to_xyxy(x):
    x_c, y_c, w, h = x.unbind(-1)
    b = [(x_c - 0.5 * w), (y_c - 0.5 * h), (x_c + 0.5 * w), (y_c + 0.5 * h)]
    return torch.stack(b, dim=1)


def rescale_bboxes(out_bbox, size):
    width, height = size
    boxes = box_cxcywh_to_xyxy(out_bbox)
    boxes = boxes * torch.tensor([width, height, width, height], dtype=torch.float32)
    return boxes


def outputs_to_objects(outputs, img_size, id2label):
    m = outputs.logits.softmax(-1).max(-1)
    pred_labels = list(m.indices.detach().cpu().numpy())[0]
    pred_scores = list(m.values.detach().cpu().numpy())[0]
    pred_bboxes = outputs['pred_boxes'].detach().cpu()[0]
    pred_bboxes = [elem.tolist() for elem in rescale_bboxes(pred_bboxes, img_size)]

    objects = []
    for label, score, bbox in zip(pred_labels, pred_scores, pred_bboxes):
        class_label = id2label[int(label)]
        if not class_label == 'no object':
            objects.append({'label': class_label, 'score': float(score),
                            'bbox': [float(elem) for elem in bbox]})

    return objects

def recognize_table(image):
    # prepare image for the model
    # pixel_values = structure_processor(images=image, return_tensors="pt").pixel_values
    pixel_values = structure_transform(image).unsqueeze(0).to(device)

    # forward pass
    with torch.no_grad():
        outputs = structure_model(pixel_values)

    # postprocess to get individual elements
    id2label = structure_model.config.id2label
    id2label[len(structure_model.config.id2label)] = "no object"
    cells = outputs_to_objects(outputs, image.size, id2label)

    # visualize cells on cropped table
    # draw = ImageDraw.Draw(image)

    # for cell in cells:
    #     draw.rectangle(cell["bbox"], outline="red")
        
    # return image, cells
    return cells



def process_pdf(image):
    cropped_table = image

    cells = recognize_table(cropped_table)
   
    cell_coordinates = get_cell_coordinates_by_row(cells)

    # df, data = apply_ocr(cell_coordinates, image)

    # return image, df, data
    
    return cells


def intersect_boxes(boxA, boxB):
    """
    Compute the intersection of two bounding boxes.
    Each box is defined as [x, y, xmax, ymax].
    """
    # xA = max(boxA[0], boxB[0])
    # yA = max(boxA[1], boxB[1])
    # xB = min(boxA[2],boxB[2])
    # yB = min(boxA[3],boxB[3])
    xA = boxB[0]
    yA = boxA[1]
    xB = boxB[2]
    yB = boxA[3]

    # Compute the width and height of the intersection box
    width = xB - xA
    height = yB - yA

    if width <= 0 or height <= 0:
        return None  # No intersection

    return [xA, yA, xB, yB]

def calculate_iou(cell,span):
    cx1,cy1,cx2,cy2 = cell
    sx1,sy1,sx2,sy2 = span

    w1 = cx2-cx1
    h1 = cy2 - cy1

    w2 = sx2 - sx1
    h2 = sy2 - sy1
    area_cell = w1*h1
    area2 = w2*h2
    
    xmin_intersect = max(cx1,sx1)
    ymin_intersect = max(cy1,sy1)
    
    xmax_intersect = min(cx2,sx2)
    ymax_intersect = min(cy2,sy2)
    
    width_intersect = max(0.0,xmax_intersect - xmin_intersect)
    height_intersect = max(0.0, ymax_intersect - ymin_intersect)

    area_intersect = width_intersect * height_intersect
    iou = area_intersect/ (area_cell)
    
    return iou



def extract_text(image_path):
     
    result = ocr.ocr(image_path, cls=True)
    text = ""
    for idx in range(len(result)):
        res = result[idx]
        if res != None :
            for line in res:
                text = line[1][0]
               
    return text

def index_to_excel_cell(index,len_columns):
    val = index%len_columns+1
    row_num = int(index/len_columns)+1
    column_letter = get_column_letter(val)
    return f"{column_letter}{row_num}"  # Assuming row 1, adjust as needed

def final_output(image_PIL):
  
    image_inf = process_pdf(image_PIL)
    
    # creating rows and columsn lists
    # image_PIL.show()
    rows=[]
    columns=[]
    for i in image_inf:
        if i['label']=='table column':
            columns.append(i['bbox'])
        elif i['label']=='table row':
            rows.append(i['bbox'])

    # extracting each cell from rows and colums
    cells = []
    for row in rows:
        for column in columns:
            cell = intersect_boxes(row, column)
            if cell:  # If there's an intersection
                cells.append(cell)

    # sorting the cells array first according to y coordinate then x
    lists = sorted(cells,key = lambda x : (x[1],x[0]))

    spans = []
    for i in image_inf:
        if i['label'] == 'table spanning cell':
            spans.append(i['bbox'])

    merge_indices = []
    
    # to map the cells to their spanned cell


    for i in spans:
        merge_pairs = []
        for index,cell in enumerate(lists):
            val = calculate_iou(cell,i)
            if val>0.5:
                merge_pairs.append(index)
        merge_indices.append(merge_pairs)



    large_image = image_PIL


    count = 0
    wb = Workbook()
    ws = wb.active
    len_columns = len(columns)
    
    for i in lists:
        x1,y1,x2,y2 = i
        cell_image = large_image.crop((x1,y1,x2, y2 ))
        text = pytesseract.image_to_string(cell_image,config='--psm 7')
        cell_address = index_to_excel_cell(count,len_columns)
        clean_text = re.sub(r'[^\x20-\x7E()%-]','', text)
        ws[cell_address] = clean_text
        count +=1

    # merging finally
    for i in merge_indices:
        if len(i)>1 :
            cell_address1 = index_to_excel_cell(min(i),len_columns)
            cell_address2 = index_to_excel_cell(max(i),len_columns)
            ws.merge_cells(f"{cell_address1}:{cell_address2}")

    count = 0


    for i in spans:

        x1,y1,x2,y2 = i
        cell_image = large_image.crop((x1,y1,x2, y2 ))

        text = pytesseract.image_to_string(cell_image,config='--psm 7')
        cell_address = index_to_excel_cell(merge_indices[count][0],len_columns)
        clean_text = re.sub(r'[^\x20-\x7E()%-]','', text)
        ws[cell_address] = clean_text
        count +=1

    wb.save("output_final.xlsx")
    



if __name__ =="__main__":
    app.run()
